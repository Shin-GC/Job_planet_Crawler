import JobPlanetCrawler as jp
from openpyxl import Workbook
from openpyxl.styles import Font, Fill,Alignment,PatternFill 
wb=Workbook()
ws=wb.active


#검색한 공고의 최대 페이지
blue_fill=PatternFill(start_color='1E90FF',end_color='1E90FF',fill_type='solid')
#파란색으로 채우기
def cell_style():
    ws['A1'] = 'Title'
    ws['A1'].fill=blue_fill
    ws['A1'].font =Font(bold = True)
    ws['B1'] = 'Company'
    ws['B1'].fill=blue_fill
    ws['B1'].font =Font(bold = True)
    ws['C1'] = 'Score'
    ws['C1'].fill=blue_fill
    ws['C1'].font =Font(bold = True)
#맨 위 구분 셀을 파란색으로 체워주기
cell_style()

titles=[]
companys=[]
scores=[]
title=[]
company=[]
score=[]

maxpage=jp.get_max_page(jp.get_soup(search='python'))
for i in range(1,maxpage):
    save=jp.get_Job_Planet(jp.get_soup(search='python',page=i))
    titles.append(save['title'])
    companys.append(save['company'])
    scores.append(save['score'])
    #title=jp.get_Job_Planet(jp.get_soup(start_star=3))['title']
    #company=jp.get_Job_Planet(jp.get_soup(start_star=3))['company']
    #score=jp.get_Job_Planet(jp.get_soup(start_star=3))['score']
title=sum(titles,title)
company=sum(companys,company)
score=sum(scores,score)

for t in range(len(title)):
    ws.append([title[t]])
    #엑셀 title A셀에 값 넣기
for c in range(2,len(company)+2):
    ws[f"B{c}"]=company[c-2]
    #엑셀 company 값 B 셀에 넣기
for s in range(2,len(score)+2):
    ws[f"C{s}"]=score[s-2]
    #엑셀 score 값 C 셀에 넣기


for x in range(1,len(title)+2):
    for y in range(1,len(title)+2):
        cell=ws.cell(row=x,column=y)
        cell.alignment=Alignment(horizontal='center')
# 가운데 정렬 [제목이 가장 기므로 제목을 기준으로 반복 길이를 정함.]

int_title=[]
int_company=[]
int_score=[]

for i in range(len(title)):
    int_title.append(len(title[i]))

for i in range(len(company)):
    int_company.append(len(company[i]))

for i in range(len(score)):
    int_score.append(len(score[i]))
# 너비를 정해주기 위해 글씨의 길이를 리스트에 저장
max_a_width=max(int_title)
max_b_width=max(int_company)
max_c_width=max(int_score)
# 저장한 글씨 길이 리스트 중 가장 긴 것을 지정
ws.column_dimensions['A'].width = max_a_width+15
ws.column_dimensions['B'].width = max_b_width*2
ws.column_dimensions['C'].width = max_c_width+10
# 최대 길이를 잰 후 엑셀 너비 지정 [ 폰트 크기를 위해 뒤에 숫자로 추가너비 기입 ]

wb.save('JobPlanet.xlsx')
wb.close()